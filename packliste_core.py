#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
import json
import math
import shutil
import datetime
from pathlib import Path

from copy import copy

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# ------------------------------------------------------------
# Konfiguration & Konstanten
# ------------------------------------------------------------

TEMPLATE_FILE = "Packliste_Template.xlsx"
DICHTUNGEN_CONFIG = "dichtungen.json"

SERVICE_TECHNIKER_ROW = 1
DATE_ROW = 2

TEMPLATE_SUM_ROW = 1
TEMPLATE_DICHTUNG_NAME_ROW = 2
TEMPLATE_DATA_START_ROW = 3

DF_DATA_START_ROW = 1
DF_SUM_ROW = 0

PLATZHALTER_COL_INDEX = 5  # Spalte E im Template
NUMBERING_COL = 1          # Spalte A

weekday_map = {
    0: "MO",
    1: "DI",
    2: "MI",
    3: "DO",
    4: "FR",
    5: "SA",
    6: "SO",
}

# Default-Liste, falls noch keine JSON vorhanden ist
DEFAULT_DICHTUNGEN = []


def resource_path(relative_path: str) -> str:
    """
    Liefert einen Pfad relativ zu dieser Datei (funktioniert auch auf dem Server).

    Für die Dichtungs-Konfiguration erlauben wir auch den Umgebungs-Parameter
    ``DICHTUNGEN_PATH``. Wenn der nicht gesetzt ist, fällt der Pfad auf die
    Projekt-Root zurück (neben ``packliste_core.py``). So können Deployments
    einen beschreibbaren Speicherort hinterlegen.
    """

    env_path = os.getenv("DICHTUNGEN_PATH")
    if env_path and Path(env_path).parent.exists():
        return env_path

    base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, relative_path)


# ------------------------------------------------------------
# Helper-Funktionen für Daten / Datumswerte
# ------------------------------------------------------------

def transform_zeitraum(val):
    """
    Wandelt '24.11.2025 08:00 - 09:00' in 'MO 24.11.25 08:00 - 09:00' um.
    """
    if not val or not isinstance(val, str):
        return val
    m = re.match(r'^(\d{1,2}\.\d{1,2}\.\d{4})(.*)$', val.strip())
    if not m:
        return val
    date_str = m.group(1).strip()
    rest = m.group(2)
    try:
        dt = datetime.datetime.strptime(date_str, "%d.%m.%Y")
        wday = dt.weekday()
        wday_abbr = weekday_map.get(wday, "")
        date_formatted = dt.strftime("%d.%m.%y")
        return f"{wday_abbr} {date_formatted}{rest}"
    except Exception:
        return val


def safe_val(df, col, index):
    """
    Sichere DataFrame-Zugriffe, damit bei fehlenden Spalten/Zeilen kein Fehler entsteht.
    """
    if col not in df.columns:
        return ""
    if index < 0 or index >= len(df):
        return ""
    val = df[col].iloc[index]
    if pd.isna(val):
        return ""
    return str(val)


def parse_number(s):
    try:
        return float(str(s).replace(",", "."))
    except Exception:
        return 0.0


def parse_date_part(value):
    """
    Holt aus einem String ein Datum im Format dd.mm.yyyy, z.B. '24.11.2025 08:00 - 09:00'.
    """
    if not value or not isinstance(value, str):
        return None
    match = re.match(r"^(\d{1,2}\.\d{1,2}\.\d{4})", value.strip())
    if not match:
        return None
    date_str = match.group(1)
    try:
        dt = pd.to_datetime(date_str, dayfirst=True, errors="coerce")
        return dt
    except Exception:
        return None


def get_zeitraum_von_bis(df, col="Zeitraum"):
    """
    Ermittelt den Gesamtzeitraum 'von - bis' aus der Zeitraum-Spalte.
    """
    if col not in df.columns:
        return ""
    dtlist = []
    for val in df[col].dropna():
        dt = parse_date_part(str(val))
        if dt is not None:
            dtlist.append(dt)
    if not dtlist:
        return ""
    von_dt = min(dtlist)
    bis_dt = max(dtlist)
    return f"{von_dt.strftime('%d.%m.%Y')} - {bis_dt.strftime('%d.%m.%Y')}"


def spalte_leer(df, colname):
    """
    True, wenn die Spalte komplett leer ist (oder gar nicht existiert).
    """
    if colname not in df.columns:
        return True
    col_series = df[colname].dropna().astype(str).str.strip()
    if len(col_series) == 0:
        return True
    return col_series.eq("").all()


# ------------------------------------------------------------
# Sortierlogik für Dichtungen
# ------------------------------------------------------------

def parse_numeric_part(name: str) -> float:
    """
    Extrahiert einen numerischen Teil, z.B. '10/5_S' -> 10.005, für sortierbare Reihenfolge.
    """
    m = re.search(r'(\d+)(?:/(\d+))?', name)
    if m:
        first = int(m.group(1))
        second = int(m.group(2)) if m.group(2) else 0
        return first + second / 1000.0
    return 999999.0


def parse_suffix_priority(name: str) -> int:
    """
    Sortier-Priorität nach Suffix: _S vor _W vor _G vor Rest.
    """
    sfx = ""
    parts = name.rsplit("_", 1)
    if len(parts) == 2 and parts[1].strip():
        sfx = parts[1].strip().upper()
    if sfx == "S":
        return 0
    elif sfx == "W":
        return 1
    elif sfx == "G":
        return 2
    else:
        return 99


def final_sort_dichtungen(dichtungen, df=None):
    """
    Sortiert die Dichtungen in sinnvoller Reihenfolge.
    Standard-Dichtungen (always_show=True) zuerst.
    """

    def sort_key(d):
        is_std = d.get("always_show", False)
        order_str = str(d.get("order", "")).strip()
        try:
            order_val = int(order_str)
        except ValueError:
            order_val = None
        name = d["name"]
        suffix_prio = parse_suffix_priority(name)
        numeric_val = parse_numeric_part(name)
        alpha_name = name.lower()

        if is_std:
            group = 0
            if order_val is not None:
                return (group, 0, order_val)
            else:
                return (group, 1, numeric_val, alpha_name)
        else:
            group = 1
            return (group, suffix_prio, numeric_val, alpha_name)

    return sorted(dichtungen, key=sort_key)


# ------------------------------------------------------------
# Helper zum Kopieren von Formatierungen
# ------------------------------------------------------------

def copy_cell_style(src_cell, dst_cell):
    if src_cell.font:
        dst_cell.font = copy(src_cell.font)
    if src_cell.border:
        dst_cell.border = copy(src_cell.border)
    if src_cell.fill:
        dst_cell.fill = copy(src_cell.fill)
    if src_cell.number_format:
        dst_cell.number_format = src_cell.number_format
    if src_cell.protection:
        dst_cell.protection = copy(src_cell.protection)
    if src_cell.alignment:
        align = copy(src_cell.alignment)
        align.wrap_text = True
        dst_cell.alignment = align
    else:
        dst_cell.alignment = Alignment(wrap_text=True)


def copy_entire_row_format(ws, src_row_idx, dst_row_idx):
    """
    Kopiert Formatierungen einer Zeile (inkl. Höhe) und leert die Inhalte.
    """
    ws.row_dimensions[dst_row_idx].height = None
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        sc = ws.cell(row=src_row_idx, column=col)
        dc = ws.cell(row=dst_row_idx, column=col)
        copy_cell_style(sc, dc)
        dc.value = None


def copy_column_with_style(ws, src_col_idx, dst_col_idx):
    """
    Kopiert eine komplette Spalte inkl. Formatierungen und Breite.
    """
    max_row = ws.max_row
    for row in range(1, max_row + 1):
        sc = ws.cell(row=row, column=src_col_idx)
        dc = ws.cell(row=row, column=dst_col_idx)
        copy_cell_style(sc, dc)
        dc.value = sc.value
    src_letter = get_column_letter(src_col_idx)
    dst_letter = get_column_letter(dst_col_idx)
    if ws.column_dimensions[src_letter].width:
        ws.column_dimensions[dst_letter].width = ws.column_dimensions[src_letter].width


def set_horizontal_dotted(ws, row_idx):
    dotted_side = Side(style="dotted", color="999999")
    max_col = ws.max_column
    for col_idx in range(1, max_col + 1):
        c = ws.cell(row=row_idx, column=col_idx)
        b = copy(c.border) or Border()
        b.top = dotted_side
        b.bottom = dotted_side
        c.border = b


def set_bottom_thick(ws, row_idx):
    thick_side = Side(style="medium")
    max_col = ws.max_column
    for col_idx in range(1, max_col + 1):
        c = ws.cell(row=row_idx, column=col_idx)
        b = copy(c.border) or Border()
        b.bottom = thick_side
        c.border = b


def set_top_border_solid(ws, row_idx):
    thin_side = Side(style="thin")
    max_col = ws.max_column
    for col_idx in range(1, max_col + 1):
        c = ws.cell(row=row_idx, column=col_idx)
        b = copy(c.border) or Border()
        b.top = thin_side
        c.border = b


def set_column_left_border(ws, col_idx, start_row=3, border_style="thin"):
    side = Side(style=border_style, color="000000")
    for row in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        b = copy(cell.border) or Border()
        b.left = side
        cell.border = b


def set_bottom_solid(ws, row_idx):
    side = Side(style="thin", color="000000")
    max_col = ws.max_column
    for col_idx in range(1, max_col + 1):
        c = ws.cell(row=row_idx, column=col_idx)
        b = copy(c.border) or Border()
        b.bottom = side
        c.border = b


def remove_trailing_blank_rows(ws, start_row):
    """
    Löscht am Tabellenende komplett leere Zeilen.
    """

    def row_is_blank(r):
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=col_idx).value
            if val not in (None, ""):
                return False
        return True

    last_row = ws.max_row
    while last_row > start_row:
        if row_is_blank(last_row):
            ws.delete_rows(last_row, 1)
        else:
            break
        last_row = ws.max_row


# ------------------------------------------------------------
# Dichtungs-Namen umbrechen & Spaltenbreiten
# ------------------------------------------------------------

def apply_dicht_name_break(name: str) -> str:
    """
    Versucht, Dichtungsnamen sinnvoll in max. 2 Zeilen umzubrechen.

    - Wenn '_' vorhanden: Umbruch dort (z.B. '10/5_S' -> '10/5\nS').
    - Sonst, wenn Leerzeichen vorhanden: Umbruch zwischen Wörtern
      (z.B. 'Omega klebend' -> 'Omega\nklebend').
    - Sonst: String ungefähr in der Mitte trennen.
    """
    if not name:
        return ""
    name = str(name).strip()

    if "_" in name:
        left, right = name.split("_", 1)
        return left + "\n" + right

    parts = name.split()
    if len(parts) == 1:
        if len(name) <= 8:
            return name
        half = math.ceil(len(name) / 2)
        return name[:half] + "\n" + name[half:]
    else:
        if len(parts) == 2:
            return parts[0] + "\n" + parts[1]
        mid = math.ceil(len(parts) / 2)
        line1 = " ".join(parts[:mid])
        line2 = " ".join(parts[mid:])
        return line1 + "\n" + line2


def adjust_dichtung_column_widths(ws, dicht_col_map, max_width=12, min_width=6):
    """
    Setzt die Spaltenbreite der Dichtungs-Spalten so klein wie möglich,
    ohne dass die Überschrift unsinnig oft umbrechen muss.
    Orientiert sich an der Länge der längeren Zeile der Überschrift.
    """
    for name, col_idx in dicht_col_map.items():
        header_cell = ws.cell(row=TEMPLATE_DICHTUNG_NAME_ROW, column=col_idx)
        txt = str(header_cell.value) if header_cell.value is not None else ""
        lines = txt.split("\n")
        max_len = max((len(line) for line in lines), default=0)
        width = max(min_width, min(max_width, max_len + 2))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width


# ------------------------------------------------------------
# Laden / Speichern der Dichtungen
# ------------------------------------------------------------

def save_dichtungen(dichtungen_list) -> bool:
    """
    Speichert die Dichtungen im JSON-Format.
    Wird von der Web-Oberfläche aufgerufen.

    Rückgabe: ``True`` bei Erfolg, sonst ``False``.
    """

    path = Path(resource_path(DICHTUNGEN_CONFIG))
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path = path.with_suffix(path.suffix + ".tmp")
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(dichtungen_list, f, ensure_ascii=False, indent=2)
        tmp_path.replace(path)
        return True
    except Exception as e:
        print("Fehler beim Speichern der Dichtungen:", e)
        return False


def load_dichtungen():
    """
    Lädt die Dichtungen aus 'dichtungen.json'.
    Falls die Datei fehlt oder fehlerhaft ist, wird eine leere Liste zurückgegeben.
    """
    path = resource_path(DICHTUNGEN_CONFIG)
    if not os.path.exists(path):
        return DEFAULT_DICHTUNGEN.copy()
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        print("Fehler beim Laden der Dichtungen:", e)
        return DEFAULT_DICHTUNGEN.copy()
    normalized = []
    for item in data:
        if isinstance(item, dict):
            normalized.append(item)
        else:
            normalized.append(
                {"name": item, "always_show": False, "default_value": 0, "order": ""}
            )
    return normalized


def guess_dichtungen_from_df(df):
    """
    Fallback: Wenn keine Dichtungen aus JSON kommen,
    nehmen wir alle Spalten außer den bekannten Feldern, die nicht komplett leer sind.
    """
    known = {
        "Service Techniker",
        "Zeitraum",
        "Dealname",
        "Weitere Techniker",
        "Informationen Packliste",
        "Ersatzteil und Zubehör",
    }
    candidates = []
    for col in df.columns:
        if col in known:
            continue
        if spalte_leer(df, col):
            continue
        candidates.append({
            "name": col,
            "always_show": True,
            "default_value": 0,
            "order": ""
        })
    return candidates


# ------------------------------------------------------------
# Hauptfunktion: Konvertierung
# ------------------------------------------------------------

def convert_file(input_path, output_path, user_dichtungen=None, show_message=False):
    """
    Konvertiert die Export-Datei (Excel/CSV) in die Packlisten-Vorlage.
    """
    # 1) Template einlesen
    template_path = resource_path(TEMPLATE_FILE)
    if not os.path.isfile(template_path):
        raise FileNotFoundError(f"Template-Datei '{TEMPLATE_FILE}' wurde nicht gefunden.")
    template_orig_wb = load_workbook(template_path)
    template_orig_ws = template_orig_wb.active
    original_width_info = template_orig_ws.column_dimensions["F"].width
    original_width_ersatz = template_orig_ws.column_dimensions["G"].width

    # 2) Eingabedatei in DataFrame laden
    ext = os.path.splitext(input_path)[1].lower()
    if ext == ".csv":
        df = pd.read_csv(input_path, sep=";", engine="python", header=0)
    else:
        df = pd.read_excel(input_path, header=0)

    # 3) Datensätze nach Datum/Uhrzeit sortieren (Zeile 0 bleibt Summenzeile)
    try:
        sum_row = df.iloc[[0]].copy()
        data_rows = df.iloc[1:].copy()
        if "Zeitraum" in data_rows.columns:

            def parse_datetime(x):
                pattern = r'^(\d{1,2}\.\d{1,2}\.\d{4})\s+(\d{1,2}:\d{1,2})'
                m = re.match(pattern, str(x))
                if m:
                    dt_str = f"{m.group(1)} {m.group(2)}"
                    return pd.to_datetime(dt_str, format="%d.%m.%Y %H:%M", errors="coerce")
                m2 = re.match(r'^(\d{1,2}\.\d{1,2}\.\d{4})', str(x))
                if m2:
                    return pd.to_datetime(m2.group(1), dayfirst=True, errors="coerce")
                return pd.NaT

            data_rows["ParsedDateTime"] = data_rows["Zeitraum"].apply(parse_datetime)
            data_rows.sort_values(by="ParsedDateTime", ascending=True, inplace=True)
            df = pd.concat([sum_row, data_rows], ignore_index=True)
            df.drop(columns=["ParsedDateTime"], inplace=True, errors="ignore")
    except Exception as e:
        print("Fehler beim Sortieren nach Datum/Uhrzeit:", e)

    # 4) Dichtungen laden bzw. erraten
    if user_dichtungen is None:
        user_dichtungen = load_dichtungen()
    if not user_dichtungen:
        user_dichtungen = guess_dichtungen_from_df(df)

    def has_effective_dichtungen(dichtungen):
        for d in dichtungen:
            name = d.get("name")
            if not name:
                continue
            if str(name).strip().lower() == "tag":
                continue
            is_standard = d.get("always_show", False)
            if is_standard:
                return True
            if name in df.columns and not spalte_leer(df, name):
                return True
        return False

    if not has_effective_dichtungen(user_dichtungen):
        user_dichtungen = guess_dichtungen_from_df(df)

    # 5) Template-Kopie erzeugen
    temp_copy_path = output_path + "_temp_template.xlsx"
    shutil.copyfile(template_path, temp_copy_path)

    wb = load_workbook(temp_copy_path)
    ws = wb.active
    ws.delete_rows(1)  # erste Zeile im Template entfernen

    # 6) Dichtungen sortieren
    final_dichtungen = final_sort_dichtungen(user_dichtungen, df)

    # 7) Kopfbereich (Technikername & Zeitraum)
    serv_val = safe_val(df, "Service Techniker", 3)
    ws.cell(row=SERVICE_TECHNIKER_ROW, column=2, value=serv_val).font = Font(
        name="Calibri", size=14, bold=True
    )
    zr = get_zeitraum_von_bis(df, "Zeitraum")
    ws.cell(row=DATE_ROW, column=2, value=zr).font = Font(
        name="Calibri", size=14, bold=True
    )

    # 8) Mapping Eingabespalten -> Template-Spalten
    global_mainfield = [
        ("Zeitraum", 2),                # B
        ("Dealname", 3),                # C
        ("Weitere Techniker", 4),       # D
        ("Informationen Packliste", 6), # F
        ("Ersatzteil und Zubehör", 7),  # G
    ]

    # 9) Dichtungs-Spalten ab PLATZHALTER_COL_INDEX (E)
    current_col = PLATZHALTER_COL_INDEX
    first_run = True
    dicht_col_map = {}

    for dicht in final_dichtungen:
        name = dicht.get("name")
        if not name:
            continue
        # "Tag" nicht als Dichtung anzeigen
        if str(name).strip().lower() == "tag":
            continue
        is_standard = dicht.get("always_show", False)
        # Nicht-Standard-Dichtungen nur anzeigen, wenn Werte vorhanden
        if (not is_standard) and spalte_leer(df, name):
            continue

        if first_run:
            used_col = current_col
            first_run = False
        else:
            new_col = current_col + 1
            ws.insert_cols(new_col)
            copy_column_with_style(ws, PLATZHALTER_COL_INDEX, new_col)
            # Mapping rechts verschieben
            for i, (dfcol, cidx) in enumerate(global_mainfield):
                if cidx >= new_col:
                    global_mainfield[i] = (dfcol, cidx + 1)
            used_col = new_col
            current_col = new_col

        # linke Rahmenlinie
        set_column_left_border(ws, used_col, start_row=1, border_style="thin")

        # Überschrift mit sinnvollem Zeilenumbruch
        mod_name = apply_dicht_name_break(name)
        head_cell = ws.cell(row=TEMPLATE_DICHTUNG_NAME_ROW, column=used_col, value=mod_name)
        head_cell.font = Font(name="Calibri", size=12, bold=True)
        head_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Summen-Zeile
        sum_val_raw = safe_val(df, name, DF_SUM_ROW)
        sum_num = parse_number(sum_val_raw)
        sum_cell = ws.cell(row=TEMPLATE_SUM_ROW, column=used_col, value=round(sum_num))
        sum_cell.number_format = "0"
        sum_cell.font = Font(name="Calibri", size=16, color="FF0000")
        sum_cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

        dicht_col_map[name] = used_col

    # Linie unter den Dichtungsnamen
    set_bottom_solid(ws, TEMPLATE_DICHTUNG_NAME_ROW)

    # 10) Datenzeilen übertragen
    t_row = TEMPLATE_DATA_START_ROW
    for df_row in range(DF_DATA_START_ROW, len(df)):
        if t_row > ws.max_row:
            ws.insert_rows(idx=t_row)
        copy_entire_row_format(ws, TEMPLATE_DATA_START_ROW, t_row)

        row_num = df_row
        num_cell = ws.cell(row=t_row, column=NUMBERING_COL, value=row_num)
        num_cell.font = Font(name="Calibri", size=12, bold=True)
        num_cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)

        # Hauptfelder
        for (df_col, tmplt_col) in global_mainfield:
            val = safe_val(df, df_col, df_row)
            cell = ws.cell(row=t_row, column=tmplt_col)
            if df_col == "Zeitraum":
                val = transform_zeitraum(val)
                cell.value = val
                cell.font = Font(name="Calibri", size=12, bold=True)
            else:
                cell.value = val
                if df_col in ["Informationen Packliste", "Ersatzteil und Zubehör", "Weitere Techniker"]:
                    cell.font = Font(bold=True, color="FF0000")
                else:
                    cell.font = Font(name="Calibri", size=12, bold=False, color="000000")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Dichtungswerte
        for dicht in final_dichtungen:
            name = dicht.get("name")
            if not name:
                continue
            if str(name).strip().lower() == "tag":
                continue
            is_standard = dicht.get("always_show", False)
            if (not is_standard) and spalte_leer(df, name):
                continue
            col_idx = dicht_col_map.get(name)
            if col_idx is None:
                continue
            raw_val = df[name].iloc[df_row] if name in df.columns else ""
            cell = ws.cell(row=t_row, column=col_idx)
            try:
                num_val = float(raw_val)
                cell.value = round(num_val)
                cell.number_format = "0"
            except Exception:
                cell.value = raw_val
            cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            cell.font = Font(name="Calibri", size=12, bold=False)

        # Linien & Zebra
        if t_row == TEMPLATE_DATA_START_ROW:
            set_top_border_solid(ws, t_row)
            set_horizontal_dotted(ws, t_row)
        else:
            set_horizontal_dotted(ws, t_row)

        bg_color = "DDDDDD" if (row_num % 2 == 1) else "FFFFFF"
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=t_row, column=col_idx).fill = PatternFill("solid", fgColor=bg_color)

        t_row += 1

    # 11) Zusätzliche-Dichtungen-Zeile
    extra_line_row = t_row
    ws.insert_rows(idx=extra_line_row)
    copy_entire_row_format(ws, TEMPLATE_DATA_START_ROW, extra_line_row)
    set_top_border_solid(ws, extra_line_row)
    set_bottom_thick(ws, extra_line_row)

    extra_text_cell = ws.cell(row=extra_line_row, column=3, value="zusätzliche Dichtungen")
    extra_text_cell.font = Font(bold=True)
    extra_text_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    last_used_df_row = len(df)
    bg_color = "DDDDDD" if (last_used_df_row % 2 == 1) else "FFFFFF"
    for col_idx in range(1, ws.max_column + 1):
        ws.cell(row=extra_line_row, column=col_idx).fill = PatternFill("solid", fgColor=bg_color)

    # 12) Standard-Dichtungen in der zusätzlichen Zeile vorbelegen
    for dicht in final_dichtungen:
        if not dicht.get("always_show", False):
            continue
        name = dicht.get("name")
        if not name or name not in dicht_col_map:
            continue
        col_idx = dicht_col_map[name]
        fix_value = dicht.get("default_value", 0)
        try:
            fix_value_num = float(fix_value)
        except Exception:
            fix_value_num = 0.0
        c = ws.cell(row=extra_line_row, column=col_idx, value=fix_value_num)
        c.number_format = "0"
        c.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        c.font = Font(name="Calibri", size=12, bold=False)

        old_sum = ws.cell(row=TEMPLATE_SUM_ROW, column=col_idx).value
        old_sum = old_sum if isinstance(old_sum, (int, float)) else 0
        new_sum = old_sum + fix_value_num
        s_cell = ws.cell(row=TEMPLATE_SUM_ROW, column=col_idx, value=new_sum)
        s_cell.number_format = "0"
        s_cell.font = Font(name="Calibri", size=16, bold=False, color="FF0000")
        s_cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    # 13) Info/Ersatzteil-Spaltenbreite aus Template übernehmen
    for field, orig_width in [
        ("Informationen Packliste", original_width_info),
        ("Ersatzteil und Zubehör", original_width_ersatz),
    ]:
        col_idx = next((col for df_field, col in global_mainfield if df_field == field), None)
        if col_idx is not None and orig_width:
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = orig_width

    # 14) Bestimmte Spalten ausblenden, wenn sie komplett leer sind
    for field in ["Weitere Techniker", "Informationen Packliste", "Ersatzteil und Zubehör"]:
        col_idx = next((col for (df_field, col) in global_mainfield if df_field == field), None)
        if col_idx is None:
            continue
        col_letter = get_column_letter(col_idx)
        if spalte_leer(df, field):
            ws.column_dimensions[col_letter].hidden = True
        else:
            ws.column_dimensions[col_letter].hidden = False

    # 15) Leere Zeilen am Ende entfernen
    remove_trailing_blank_rows(ws, extra_line_row)

    # 16) Schriftfarbe der Dichtungs-Spalten alternierend blau/schwarz
    BLUE_COLOR = "0000FF"
    BLACK_COLOR = "000000"
    dicht_spalten_sorted = sorted(dicht_col_map.values())
    for i, col_idx in enumerate(dicht_spalten_sorted):
        font_color = BLUE_COLOR if i % 2 == 0 else BLACK_COLOR
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            new_font = copy(cell.font)
            new_font.color = font_color
            cell.font = new_font

    # 17) Dichtungs-Spaltenbreiten anpassen
    adjust_dichtung_column_widths(ws, dicht_col_map)

    # 18) Speichern
    wb.save(output_path)
    wb.close()
    if os.path.exists(temp_copy_path):
        os.remove(temp_copy_path)

    return output_path
